from openpyxl import load_workbook


class TradeImporter:
    def __init__(self, file_name):
        self.file_name = file_name

    current_category_level = [None] * 10
    category_list = []
    product_list = []

    category_with_only_products = ['УП-00044771', '00000000006', 'УП-00069001', '6',
                                   '00000000001', 'НК000000009', 'УП-00083763']

    def process_import(self, handler_create_category=None, handler_create_product=None):
        self.current_category_level = [None] * 10
        print('LOAD XLSX FILE....')
        workbook = load_workbook(filename=self.file_name)
        sheet = workbook.active
        index = 2
        print('XLS PARSED TRY ANALISE')
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=4):
            if not index % 1000:
                print(f'processed {index} rows')
            if row[0].value is None or row[1].value is None:
                continue
            level = sheet.row_dimensions.get(index).outlineLevel
            parent = self.current_category_level[level - 1] if level > 0 else None
            if (row[2].value is None and row[3].value is None) and not (parent in self.category_with_only_products):
                current = {
                    'name': row[0].value,
                    'code': row[1].value,
                    'parent': parent,
                    'level': level,
                    'info': '[{}:{}]'.format(row[2].value, row[3].value)
                }
                if handler_create_category:
                    current = handler_create_category(current)
                if current:
                    self.category_list.append(current)
                self.current_category_level[level] = row[1].value
            else:
                product = {
                    'name': row[0].value,
                    'code': row[1].value,
                    'vendor_code': row[2].value,
                    'brand': row[3].value,
                    #'parent': self.current_category_level[level],
                    'parent': parent,
                }
                if handler_create_product:
                    product = handler_create_product(product)
                self.product_list.append(product)
            index += 1


if __name__ == '__main__':
    importer = TradeImporter('../../import.xlsx')
    importer.process_import()

    print(importer.category_list)
    print(len(importer.category_list))
    print(len(importer.product_list))
    print(importer.product_list)
